import re, os, requests
import pandas as pd
import urllib.request
import tweepy, facebook, time
from dateutil.parser import parse
from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image

#Function to share twitter credentials:
def twitter_cred():
	auth = tweepy.OAuthHandler('consumer key', 'consumer secret')
	auth.set_access_token('access token','access token secret')
	api = tweepy.API(auth)
	return auth, api

"""Should be refactored using filter and the itertools library. Tweets out text from spreadsheet, for only 
those entries categorized for twitter."""
def tweet_bulk(df,mypath):
	auth, api = twitter_cred()

	for index,row in df.iterrows():
		if df.loc[index,'socialmedia_channel'] == 'Twitter':
			img_url = row['post_img']
			if isinstance(img_url, str):
				for root, dirs, files in os.walk(mypath + '\static'):
					for name in files:
						if name in img_url:
							filename = name
							img = mypath + '\static\\' + filename							
							api.update_with_media(img, status=row['post_text'])
			else:
				api.update_status(row['post_text'])

"""This one is for use with the scheduler function. Tweet out specific post instead of all of them.
Should be merged with the previous function. """
def tweet(_index, df):
	auth, api = twitter_cred()
	dirpath = os.getcwd()
	mypath = dirpath

	#Could also be refactored to make use of filter.
	for index,row in df.iterrows():
		if df.loc[index,'socialmedia_channel'] == 'Twitter' and index == _index:
			img_url = row['post_img']
			if isinstance(img_url, str):
				for root, dirs, files in os.walk(mypath):
					for name in files:
						if name in img_url:
							filename = name
							img = mypath + '\\' + filename
							api.update_with_media(img, status=row['post_text'])
			else:
				api.update_status(row['post_text'])

#Adds the values to the socialmedia_channel column in the spreadsheet.
def update_spreadsheet(df,ws,book):
	socialmedia_list = []
	for r in dataframe_to_rows(df, index=True, header=True):
		socialmedia_list.append(r[len(r)-1])
	socialmedia_list.remove(socialmedia_list[0])
	
	for cell in ws['E'][1:-1]:
		cell.value = socialmedia_list[ws['E'].index(cell) - 1]

	book.save('Python_exercise_Handle.xlsx')

def logo():
	urllib.request.urlretrieve("http://cityread.london/wp-content/uploads/2016/02/HarperCollins-logo.png"
		, 'harper.jpg')

	harper = Image.open('harper.jpg')
	#Convert and save to RGBA
	harper = harper.crop((200, 170, 540, 550)).convert('RGBA').save('harper.png')
	harper = Image.open('harper.png')
	return harper

#Set up a schedule using APScheduler. Runs in background until you quite using CTRL-Z.
def schedule_tweets(df):
	sched = BackgroundScheduler()

	#add scheduled jobs
	for index,row in df.iterrows():
		_time = parse(str(row['post_datetime']))
		
		sched.add_job(tweet,'date', run_date=_time, args=[index, df])
								
	sched.start()
	
	try:
		while True:
			time.sleep(2)
	except(KeyboardInterrupt, SystemExit):
		sched.shutdown()

def get_images_and_merge(df,mypath,harper):
	"""Download and save images. Need to create a folder called 'static' in root dictionary.
	In the future merge with the loop below?"""
	for img_url in df['post_img']:
		if isinstance(img_url, str):
			try:
				name = list(filter(lambda x: '.jpg' in x or '.jpeg' in x, re.split("[, \-!?:/]+",img_url)))
				filename = os.path.join(mypath + '\static', name[0])
				urllib.request.urlretrieve(img_url, filename)
			except:
				print("URL not working!")
				
	"""Merge images and save. Because of the big differences in sizes, loop will resize the harper logo
	differently according to the size of the other image."""
	for file in os.listdir(mypath + '\static'):
		os.chdir(r'c:\\Users\User\code\harper-collins\static')
		i = Image.open(file)
		if i.size[1] > 1800:
			box = (0,1810)
			i.paste(harper.resize((150,150)), box, mask=harper.resize((150,150)))
		elif i.size[1] > 1550 and i.size[1] < 1700:
			box = (0,1470)
			i.paste(harper.resize((100,100)), box, mask=harper.resize((100,100)))
		elif i.size[1] > 1500 and i.size[1] < 1750:
			box = (0,1235)
			i.paste(harper.resize((300,300)), box, mask=harper.resize((300,300)))
		elif i.size[1] > 1400 and i.size[1] < 1500:
			box = (0, 1200)
			i.paste(harper.resize((200,200)), box, mask=harper.resize((200,200)))
		elif i.size[1] > 1000 and i.size[1] < 1300:
			box = (0, 1100)
			i.paste(harper.resize((200,200)), box, mask=harper.resize((200,200)))
		elif i.size[1] > 600 and i.size[1] < 800:
			box = (0, 540)
			i.paste(harper.resize((100,100)), box, mask=harper.resize((100,100)))
		elif i.size[1] > 500 and i.size[1] < 600:
			box = (0, 430)
			i.paste(harper.resize((100,100)), box, mask=harper.resize((100,100)))
		elif i.size[1] < 500:
			box = (0, 438)
			i.paste(harper.resize((60,60)),box, mask=harper.resize((60,60)))
		i.save(file)
		#Get rid of uneeded imaged files:
		for root, dirs, files in os.walk(mypath):
			for name in files:
				if name == 'harper.jpg' or name == 'harper.png':
					os.remove(os.path.join( mypath, name ))

#Main function which calls the others and executes their code, as well as setting up initial variables.
def main():
	#load necessary workbooks and set current working directory.
	book = load_workbook('Python_exercise_Handle.xlsx')
	ws = book.get_sheet_by_name("data_table")
	df = pd.read_excel('Python_exercise_Handle.xlsx', sheet_name='data_table')
	dirpath = os.getcwd()
	mypath = dirpath
	
	#Set socialmedia_channel values
	for index, row in df.iterrows():
		if len(row['post_text']) > 140:
			df.loc[index,'socialmedia_channel'] = 'Facebook'
		else:
			df.loc[index,'socialmedia_channel'] = 'Twitter'
	
	#Get socialmedia values to append onto spreadsheet
	update_spreadsheet(df,ws, book)

	#Download harpercollins logo and make background transparent.
	harper = logo()

	get_images_and_merge(df,mypath,harper)

	#Now that images have been resized and saved, Tweet out posts.
	tweet_bulk(df,mypath)
	
	#We can also schedule these tweets. Exit from the script using CTRL-Z.
	schedule_tweets(df)
	
if __name__ == '__main__':
	main()