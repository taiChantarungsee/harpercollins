from openpyxl import Workbook
from pprint import pprint
from facepy import GraphAPI

def facebook_info():
	#Open excel shet.
	book = Workbook()
	data = []

	"""Build up URL first. For saftey I've taken out the page id. Just hit the Harper Collins FB URL:
	"graph.facebook.com/v2.11/HarperCollinsPublishersUK" To obtain it, then insert it below at line 15."""
	graph = GraphAPI('EAACEdEose0cBADZB9gVNmXUR2Abd8kN2O1Wpu85IAVTIzb5SZCanNVqoynHDORxegA5RyLn7ayIQxZBJdfL2ux9EzsTT4Kr0Y4gwXsokaCZA52Gkxu1lg1GcZBkp2wwZCQNOD8nKzZB2uznjJSKq5LE34DBkN6hCLMmJnYnU7WKTKZBeEJAOHSc6HcKdVyHMFGSwS74AAe3RowZDZD')
	posts = graph.get('501102899913552/posts?fields=message,created_time,likes.summary(true),comments.summary(true),shares,type&since=2017-05-01&until=2017-11-19', pagination=True)

	#Append posts to a list.
	while True:
		try:
			for p in posts['data']:
				data.append(p)
			posts = requests.get(posts['paging']['next']).json()
		except KeyError:
			break

	#Now create an excel sheet. And append a row of header names.
	ws = book.active
	ws.title = "Facebook Data"
	for col in range(0,1):
		ws.append(["id","likes","comments","shares","interactions","created_time", "type"])
	
	"""This is where most of the work is done. The nested loops and if statements are messy, but it
	is even more fiddly doing it with dictionaries. I.e. Using a dictionary as a switch statement to
	accesss another nested dictionary, by passing a value of the data dictionary created above. In this
	case a more straightforward approach was easier to implement. So instead we simply loop through
	each column of each row, check that there are keys in the data sub-dictionary that match, and then
	simply using the key-name of the ones that do to fetch the right data."""
	titles = ["id","likes","comments","shares","interactions","created_time", "type"]	
	count = 0
	for row in range(2,len(data)):
		for col in range(1,8):
			for key in titles:
				value = ""
				if key in data[count].keys():
					if key == ws.cell(column=col,row=1).value:
						#print(key,ws.cell(column=col,row=1).value,data[count]["id"])
						if key == "id":
							value = data[count]["id"]
						elif key == "likes":
							value = data[count]["likes"]["summary"]["total_count"]
						elif key == "created_time":
							value = data[count]["created_time"]
						elif key == "comments":
							value = data[count]["comments"]["summary"]["total_count"]
						elif key == "shares":
							value = data[count]["shares"]["count"]
						elif key == "type":
							value = data[count]["type"]
						_ = ws.cell(column=col,row=row, value=value)
		count += 1

	#Calculate total number of interactions
	for row in range(2,len(data)):
		if ws.cell(column=4,row=row).value == None:
			_ = ws.cell(column=4,row=row, value=0)
		value = (ws.cell(column=2,row=row).value + 
		ws.cell(column=3,row=row).value + ws.cell(column=4,row=row).value)
		_ = ws.cell(column=5,row=row, value=value)

	book.save(filename = 'Facebook_data.xlsx')

if __name__ == '__main__':
	facebook_info()